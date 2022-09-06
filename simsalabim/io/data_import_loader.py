"""
Data Import Loader
============

Provides
  1. IO stuff to access data from messy spreadsheets
  2. Conversion of LabVIEW TDMS files
"""
import pandas as pd
from openpyxl import load_workbook
from nptdms import TdmsFile
from pathlib import Path


def load_xl_range(io, sheet_name=0, start=None, stop=None):
    """Load data from a certain range of an Excel sheet
    and transforms it to an dataframe.

    Parameters
    ----------
    io : str
        path of the excel file
    sheet_name : str
        sheet name where the data lies

    Returns
    -------
    DataFrame

    See Also
    --------
    openpyxl.load_workbook() : Open the given filename and return the workbook.
    """
    # Load sheet of workbook
    wb = load_workbook(filename=io, read_only=True)
    sheets = wb.sheetnames

    if sheet_name in sheets:
        ws = wb[sheet_name]
    else:
        try:
            ws = wb[sheets[sheet_name]]
        except:
            raise ValueError( "sheet_name '{}' not existing in Workbook!".format(sheet_name) )

    # Read the cell values into a list of lists
    data_rows = []
    for row in ws[start:stop]:
        data_cols = []
        for cell in row:
            data_cols.append(cell.value)
        data_rows.append(data_cols)

    # Transform into dataframe
    df_out = pd.DataFrame(data_rows)
    return df_out


def read_tdms_as_dict(path):
    """Create a dictionary from a LabVIEW TDMS file.
    
    Parameters
    ----------
    path : str
        path of the tdms file

    Returns
    -------
    Dictionary

    See Also
    --------
    nptdms.TdmsFile.read() : Creates a new TdmsFile object and reads all data in the file.

    Notes
    -----
    The output dictionary looks like this:


        data = {
            group1: {
                channel1
                channel2
                ...
            }
            group2: {
                channel1
                channel2
                ...
            }
            ...
        }

    Reference
    ---------
    https://www.ni.com/de-de/support/documentation/supplemental/06/the-ni-tdms-file-format.html

    
    """  
    tdms_file = TdmsFile(path)
    dict_data ={}

    # Iteration loop over all groups
    for group in tdms_file.groups():
        dict_group = {}
        group_name = group.name
        #print(f"Group: {group_name}")

        # Iteration loop over all channels
        for channel in group.channels():
            channel_name = channel.name
            properties = channel.properties
            #print(f"Channel: {channel_name}")
            data = channel[:]

            dict_group.update({channel_name: data})

        dict_data.update({group_name: dict_group})
    
    return dict_data