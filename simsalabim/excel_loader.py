"""
Excel Loader
============

Provides
  1. IO stuff to access data from messy spreadsheets.

"""
import pandas as pd
from openpyxl import load_workbook


def load_xl_range(io, sheet_name=0, start=None, stop=None):
    """Load data from a certain range of an Excel sheet
    and transforms it to an dataframe.

    Parameters
    ----------
    io : str
        path of the excel file
    sheet_name : str
        sheet name where the data lies

    Returns
    -------
    DataFrame

    See Also
    --------
    openpyxl.load_workbook() : Open the given filename and return the workbook.
    """

    wb = load_workbook(filename=io, read_only=True)

    ws = wb[sheet_name]

    # Read the cell values into a list of lists
    data_rows = []
    for row in ws[start:stop]:
        data_cols = []
        for cell in row:
            data_cols.append(cell.value)
        data_rows.append(data_cols)

    # Transform into dataframe
    df_out = pd.DataFrame(data_rows)
    return df_out
